import pandas as pd
import os
import logging
from openpyxl import load_workbook, Workbook
from models import AnalysisSession, MaterialAnalysis
from app import db

logger = logging.getLogger(__name__)

class FileProcessor:
    def __init__(self):
        self.required_columns = [
            'manufacturer', 'country_of_origin', 'hs_code', 
            'cost_per_pair', 'fob_with_tooling', 'material_name'
        ]
    
    def process_file(self, filepath):
        """Process uploaded Excel or CSV file and extract costing data"""
        try:
            file_ext = os.path.splitext(filepath)[1].lower()
            
            if file_ext in ['.xlsx', '.xls']:
                # Try to read Excel file
                df = pd.read_excel(filepath, sheet_name=0)
            elif file_ext == '.csv':
                # Try to read CSV file
                df = pd.read_csv(filepath)
            else:
                raise ValueError("Unsupported file format")
            
            # Clean column names (lowercase, replace spaces with underscores)
            df.columns = df.columns.str.lower().str.replace(' ', '_').str.replace('/', '_')
            
            # Log available columns for debugging
            logger.info(f"Available columns: {list(df.columns)}")
            
            # Try to map common column variations
            column_mapping = {
                'manufacturer': ['manufacturer', 'mfg', 'supplier', 'vendor'],
                'country_of_origin': ['country_of_origin', 'country', 'origin', 'coo'],
                'hs_code': ['hs_code', 'hscode', 'hs', 'tariff_code'],
                'cost_per_pair': ['cost_per_pair', 'unit_cost', 'cost', 'price'],
                'fob_with_tooling': ['fob_with_tooling', 'fob', 'total_cost'],
                'material_name': ['material_name', 'material', 'component', 'description']
            }
            
            # Map columns
            mapped_columns = {}
            for required_col, possible_names in column_mapping.items():
                for possible_name in possible_names:
                    if possible_name in df.columns:
                        mapped_columns[required_col] = possible_name
                        break
            
            logger.info(f"Mapped columns: {mapped_columns}")
            
            # Check if we have essential columns
            if 'manufacturer' not in mapped_columns:
                # Try to find manufacturer in first few rows if not in columns
                manufacturer = self._extract_manufacturer_from_content(df)
                if manufacturer:
                    mapped_columns['manufacturer'] = manufacturer
            
            # Rename columns to standardized names
            df_renamed = df.rename(columns={v: k for k, v in mapped_columns.items()})
            
            # Fill missing manufacturer if found as a value
            if 'manufacturer' in mapped_columns and isinstance(mapped_columns['manufacturer'], str):
                df_renamed['manufacturer'] = mapped_columns['manufacturer']
            
            # Clean and validate data
            df_clean = self._clean_data(df_renamed)
            
            return df_clean.to_dict('records')
            
        except Exception as e:
            logger.error(f"Error processing file {filepath}: {str(e)}")
            raise
    
    def _extract_manufacturer_from_content(self, df):
        """Try to extract manufacturer from cell content"""
        # Look for manufacturer in first few rows
        for i in range(min(5, len(df))):
            for col in df.columns:
                cell_value = str(df.iloc[i][col]).lower()
                if 'manufacturer' in cell_value or 'supplier' in cell_value:
                    # Try to extract the value after the label
                    parts = cell_value.split(':')
                    if len(parts) > 1:
                        return parts[1].strip()
        return None
    
    def _clean_data(self, df):
        """Clean and validate the dataframe"""
        # Remove completely empty rows
        df = df.dropna(how='all')
        
        # Convert numeric columns
        numeric_columns = ['cost_per_pair', 'fob_with_tooling']
        for col in numeric_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
        
        # Clean string columns
        string_columns = ['manufacturer', 'country_of_origin', 'hs_code', 'material_name']
        for col in string_columns:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip()
        
        # Remove rows where all important columns are NaN
        important_cols = [col for col in ['material_name', 'country_of_origin'] if col in df.columns]
        if important_cols:
            df = df.dropna(subset=important_cols, how='all')
        
        return df
    
    def generate_results_report(self, session_id):
        """Generate Excel report with analysis results"""
        session = AnalysisSession.query.get(session_id)
        materials = MaterialAnalysis.query.filter_by(session_id=session_id).all()
        
        # Create workbook
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Analysis Summary"
        
        # Add summary data
        summary_data = [
            ["File Name", session.filename],
            ["Analysis Date", session.upload_timestamp.strftime("%Y-%m-%d %H:%M")],
            ["Manufacturer", session.manufacturer or "Not Found"],
            ["Final HS Code", session.final_hs_code or "Not Found"],
            ["Final Result", session.final_result or "Incomplete"],
            ["Reason", session.result_reason or "Analysis incomplete"],
            ["Missing Fields", ", ".join(session.missing_fields or [])]
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row_idx, column=1, value=label)
            ws_summary.cell(row=row_idx, column=2, value=value)
        
        # Materials sheet
        if materials:
            ws_materials = wb.create_sheet("Material Analysis")
            
            # Headers
            headers = ["Material Name", "Country of Origin", "HS Code", 
                      "Cost per Pair", "Problematic", "Analysis Notes"]
            for col_idx, header in enumerate(headers, 1):
                ws_materials.cell(row=1, column=col_idx, value=header)
            
            # Material data
            for row_idx, material in enumerate(materials, 2):
                ws_materials.cell(row=row_idx, column=1, value=material.material_name)
                ws_materials.cell(row=row_idx, column=2, value=material.country_of_origin)
                ws_materials.cell(row=row_idx, column=3, value=material.hs_code)
                ws_materials.cell(row=row_idx, column=4, value=material.cost_per_pair)
                ws_materials.cell(row=row_idx, column=5, value="Yes" if material.is_problematic else "No")
                ws_materials.cell(row=row_idx, column=6, value=material.analysis_notes)
        
        # Save file
        results_filename = f"fta_results_{session_id}.xlsx"
        results_path = os.path.join("uploads", results_filename)
        wb.save(results_path)
        
        return results_path
